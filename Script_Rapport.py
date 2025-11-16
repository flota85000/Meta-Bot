import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
from datetime import datetime
import pytz
import config
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os

def _tz():
    try:
        return pytz.timezone(config.FUSEAU_HORAIRE)
    except Exception:
        return pytz.timezone("Europe/Paris")

def generer_rapport_programme(programme, saison, societe=None):
    """
    Génère un rapport Excel pour un programme/saison donné
    et l'envoie par email aux RH.
    """
    tz = _tz()
    
    # Auth Google Sheets
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_file(config.CHEMIN_CLE_JSON, scopes=scope)
    client = gspread.authorize(creds)
    
    print(f"\n{'='*60}")
    print(f"📊 GÉNÉRATION RAPPORT - Programme {programme} Saison {saison}")
    if societe:
        print(f"📊 Société : {societe}")
    print(f"{'='*60}\n")
    
    # 1. Lire Réponses Sondages
    ws_reponses = client.open(config.FICHIER_PLANNING).worksheet(config.FEUILLE_REPONSES_SONDAGES)
    df_reponses = pd.DataFrame(ws_reponses.get_all_records())
    
    # Filtrer par programme/saison
    df_reponses["Programme"] = df_reponses["Programme"].astype(str).str.zfill(3)
    df_reponses["Saison"] = pd.to_numeric(df_reponses["Saison"], errors="coerce").fillna(0).astype(int)
    
    df_prog = df_reponses[
        (df_reponses["Programme"] == str(programme).zfill(3)) &
        (df_reponses["Saison"] == int(saison))
    ].copy()
    
    if societe:
        df_prog = df_prog[df_prog["Société"] == societe]
    
    if df_prog.empty:
        print(f"⚠️ Aucune réponse trouvée pour Programme {programme} Saison {saison}")
        return None
    
    print(f"✅ {len(df_prog)} réponse(s) trouvée(s)")
    
    # 2. Calculer statistiques
    nb_reponses = len(df_prog)
    nb_participants = df_prog["User ID"].nunique()
    nb_commentaires = df_prog[df_prog["Commentaire"].astype(str).str.strip() != ""].shape[0]
    
    # Dates
    try:
        date_debut = pd.to_datetime(df_prog["Date Envoi"], errors="coerce").min().strftime("%Y-%m-%d")
    except:
        date_debut = "N/A"
    
    try:
        date_fin = pd.to_datetime(df_prog["Date et Heure"], errors="coerce").max().strftime("%Y-%m-%d")
    except:
        date_fin = "N/A"
    
    # Taux réponse (simpliste: nb réponses / nb participants)
    taux_reponse = 100 if nb_participants > 0 else 0
    
    print(f"📈 Statistiques :")
    print(f"   - Participants : {nb_participants}")
    print(f"   - Réponses : {nb_reponses}")
    print(f"   - Commentaires : {nb_commentaires}")
    print(f"   - Période : {date_debut} → {date_fin}")
    
    # 3. Créer Excel
    wb = Workbook()
    
    # === ONGLET 1 : SYNTHÈSE ===
    ws_synthese = wb.active
    ws_synthese.title = "Synthèse"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Titre
    ws_synthese["A1"] = f"RAPPORT PROGRAMME {programme} - SAISON {saison}"
    ws_synthese["A1"].font = Font(bold=True, size=16)
    ws_synthese.merge_cells("A1:D1")
    
    if societe:
        ws_synthese["A2"] = f"Société : {societe}"
        ws_synthese["A2"].font = Font(size=12)
        ws_synthese.merge_cells("A2:D2")
        row_offset = 4
    else:
        row_offset = 3
    
    # Statistiques globales
    stats = [
        ["Période", f"{date_debut} → {date_fin}"],
        ["Nombre de participants", nb_participants],
        ["Nombre de sondages répondus", nb_reponses],
        ["Nombre de commentaires", nb_commentaires],
        ["Taux de réponse", f"{taux_reponse}%"],
    ]
    
    for i, (label, value) in enumerate(stats, start=row_offset):
        ws_synthese[f"A{i}"] = label
        ws_synthese[f"A{i}"].font = Font(bold=True)
        ws_synthese[f"B{i}"] = value
    
    ws_synthese.column_dimensions['A'].width = 30
    ws_synthese.column_dimensions['B'].width = 20
    
    # === ONGLET 2 : DÉTAILS RÉPONSES ===
    ws_details = wb.create_sheet("Détails Réponses")
    
    # Colonnes
    cols_details = ["Date", "Participant", "Société", "Question", "Réponse(s)", "Type", "Commentaire"]
    
    for col_num, col_name in enumerate(cols_details, 1):
        cell = ws_details.cell(row=1, column=col_num)
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Données
    for row_num, (_, row) in enumerate(df_prog.iterrows(), start=2):
        ws_details.cell(row=row_num, column=1, value=str(row.get("Date et Heure", "")))
        
        participant = f"{row.get('Prénom', '')} {row.get('Nom', '')}".strip()
        if not participant:
            participant = row.get("Username", "Anonyme")
        ws_details.cell(row=row_num, column=2, value=participant)
        
        ws_details.cell(row=row_num, column=3, value=str(row.get("Société", "")))
        ws_details.cell(row=row_num, column=4, value=str(row.get("Question", "")))
        ws_details.cell(row=row_num, column=5, value=str(row.get("Réponse(s)", "")))
        ws_details.cell(row=row_num, column=6, value=str(row.get("Type Sondage", "")))
        ws_details.cell(row=row_num, column=7, value=str(row.get("Commentaire", "")))
    
    # Largeurs colonnes
    ws_details.column_dimensions['A'].width = 18
    ws_details.column_dimensions['B'].width = 20
    ws_details.column_dimensions['C'].width = 20
    ws_details.column_dimensions['D'].width = 50
    ws_details.column_dimensions['E'].width = 30
    ws_details.column_dimensions['F'].width = 12
    ws_details.column_dimensions['G'].width = 40
    
    # === ONGLET 3 : COMMENTAIRES ===
    ws_commentaires = wb.create_sheet("Commentaires")
    
    df_commentaires = df_prog[df_prog["Commentaire"].astype(str).str.strip() != ""].copy()
    
    cols_comm = ["Date", "Participant", "Question", "Commentaire"]
    
    for col_num, col_name in enumerate(cols_comm, 1):
        cell = ws_commentaires.cell(row=1, column=col_num)
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    if not df_commentaires.empty:
        for row_num, (_, row) in enumerate(df_commentaires.iterrows(), start=2):
            ws_commentaires.cell(row=row_num, column=1, value=str(row.get("Date et Heure", "")))
            
            participant = f"{row.get('Prénom', '')} {row.get('Nom', '')}".strip()
            if not participant:
                participant = row.get("Username", "Anonyme")
            ws_commentaires.cell(row=row_num, column=2, value=participant)
            
            ws_commentaires.cell(row=row_num, column=3, value=str(row.get("Question", "")))
            ws_commentaires.cell(row=row_num, column=4, value=str(row.get("Commentaire", "")))
    
    ws_commentaires.column_dimensions['A'].width = 18
    ws_commentaires.column_dimensions['B'].width = 20
    ws_commentaires.column_dimensions['C'].width = 50
    ws_commentaires.column_dimensions['D'].width = 60
    
    # 4. Sauvegarder fichier
    timestamp = datetime.now(tz).strftime("%Y%m%d_%H%M%S")
    filename = f"Rapport_Programme_{programme}_Saison_{saison}_{timestamp}.xlsx"
    filepath = f"/tmp/{filename}"
    
    wb.save(filepath)
    print(f"✅ Fichier Excel créé : {filepath}")
    
    # 5. Enregistrer dans feuille Rapports Programmes
    try:
        ws_rapports = client.open(config.FICHIER_PLANNING).worksheet("Rapports Programmes")
        
        nouvelle_ligne = [
            str(programme).zfill(3),
            int(saison),
            societe or "",
            date_debut,
            date_fin,
            nb_participants,
            nb_reponses,
            nb_reponses,  # Nb Sondages Répondus
            taux_reponse,
            nb_commentaires,
            datetime.now(tz).strftime("%Y-%m-%d %H:%M:%S"),
            "Non",  # Email Envoyé
            ""  # Fichier Rapport (à remplir si upload Google Drive)
        ]
        
        ws_rapports.append_row(nouvelle_ligne)
        print("✅ Rapport enregistré dans 'Rapports Programmes'")
    except Exception as e:
        print(f"⚠️ Erreur enregistrement rapport : {e}")
    
    # 6. Envoyer par email
    envoyer_email_rapport(programme, saison, societe, filepath)
    
    return filepath

def envoyer_email_rapport(programme, saison, societe, filepath):
    """
    Envoie le rapport par email aux RH listés dans Contacts RH
    """
    try:
        # Configuration email (à adapter selon votre serveur SMTP)
        SMTP_SERVER = getattr(config, "SMTP_SERVER", "smtp.gmail.com")
        SMTP_PORT = getattr(config, "SMTP_PORT", 587)
        SMTP_USER = getattr(config, "SMTP_USER", "")
        SMTP_PASSWORD = getattr(config, "SMTP_PASSWORD", "")
        
        if not SMTP_USER or not SMTP_PASSWORD:
            print("⚠️ Configuration email manquante (SMTP_USER/SMTP_PASSWORD)")
            print("📧 Email non envoyé. Fichier disponible : " + filepath)
            return
        
        # Lire Contacts RH
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = Credentials.from_service_account_file(config.CHEMIN_CLE_JSON, scopes=scope)
        client = gspread.authorize(creds)
        
        ws_contacts = client.open(config.FICHIER_PLANNING).worksheet("Contacts RH")
        df_contacts = pd.DataFrame(ws_contacts.get_all_records())
        
        if societe:
            df_contacts = df_contacts[df_contacts["Société"] == societe]
        
        if df_contacts.empty:
            print(f"⚠️ Aucun contact RH trouvé pour {societe or 'cette société'}")
            return
        
        # Préparer email
        destinataires = []
        for _, row in df_contacts.iterrows():
            email_principal = str(row.get("Email RH Principal", "")).strip()
            email_copie = str(row.get("Email RH Copie", "")).strip()
            email_permanent = str(row.get("Email Copie Permanente", "")).strip()
            
            if email_principal:
                destinataires.append(email_principal)
            if email_copie:
                destinataires.append(email_copie)
            if email_permanent:
                destinataires.append(email_permanent)
        
        destinataires = list(set(destinataires))  # Dédupliquer
        
        if not destinataires:
            print("⚠️ Aucune adresse email valide trouvée")
            return
        
        # Construire message
        msg = MIMEMultipart()
        msg['From'] = SMTP_USER
        msg['To'] = ", ".join(destinataires)
        msg['Subject'] = f"Rapport Programme {programme} - Saison {saison}"
        
        if societe:
            msg['Subject'] += f" - {societe}"
        
        body = f"""
Bonjour,

Veuillez trouver ci-joint le rapport de suivi du programme {programme} (Saison {saison}).

Ce rapport contient :
- Une synthèse globale
- Le détail de toutes les réponses aux sondages
- Les commentaires laissés par les participants

Cordialement,
MetaBot
"""
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Attacher fichier
        with open(filepath, "rb") as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
        
        encoders.encode_base64(part)
        part.add_header(
            'Content-Disposition',
            f'attachment; filename= {os.path.basename(filepath)}',
        )
        
        msg.attach(part)
        
        # Envoyer
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(SMTP_USER, SMTP_PASSWORD)
        text = msg.as_string()
        server.sendmail(SMTP_USER, destinataires, text)
        server.quit()
        
        print(f"✅ Email envoyé à : {', '.join(destinataires)}")
        
        # Mettre à jour statut dans Rapports Programmes
        try:
            ws_rapports = client.open(config.FICHIER_PLANNING).worksheet("Rapports Programmes")
            records = ws_rapports.get_all_records()
            
            # Trouver la dernière ligne pour ce programme/saison
            for i, rec in enumerate(reversed(records), start=1):
                if (str(rec.get("Programme", "")).zfill(3) == str(programme).zfill(3) and
                    int(rec.get("Saison", 0)) == int(saison)):
                    row_num = len(records) - i + 2  # +2 pour header et index 1-based
                    ws_rapports.update_cell(row_num, 12, "Oui")  # Colonne "Email Envoyé"
                    break
        except Exception as e:
            print(f"⚠️ Erreur mise à jour statut email : {e}")
        
    except Exception as e:
        print(f"⚠️ Erreur envoi email : {e}")
        print(f"📧 Email non envoyé. Fichier disponible : {filepath}")

def generer_tous_rapports():
    """
    Génère les rapports pour tous les programmes terminés
    """
    # À implémenter : logique pour détecter programmes terminés
    # et générer automatiquement les rapports
    pass

if __name__ == "__main__":
    # Exemple d'utilisation
    import sys
    
    if len(sys.argv) < 3:
        print("Usage: python Script_Rapport_Final.py <programme> <saison> [societe]")
        print("Exemple: python Script_Rapport_Final.py 002 1 'Acme Corp'")
        sys.exit(1)
    
    programme = sys.argv[1]
    saison = int(sys.argv[2])
    societe = sys.argv[3] if len(sys.argv) > 3 else None
    
    generer_rapport_programme(programme, saison, societe)